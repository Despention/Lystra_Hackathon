import io
import logging

import openpyxl
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from app.database import Analysis, Issue, async_session
from app.services.export_xlsx import export_history, export_single_analysis

logger = logging.getLogger(__name__)

try:
    from xhtml2pdf import pisa
    _PDF_AVAILABLE = True
except ImportError:  # pragma: no cover
    _PDF_AVAILABLE = False
    logger.warning("xhtml2pdf not installed — PDF export will fall back to HTML")

router = APIRouter()

REPORT_TEMPLATE = """\
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>TZ Analyzer — Отчёт</title>
<style>
  body {{ font-family: 'Inter', Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 40px 20px; color: #111827; background: #fff; }}
  h1 {{ color: #2563EB; font-size: 24px; border-bottom: 2px solid #2563EB; padding-bottom: 8px; }}
  h2 {{ color: #374151; font-size: 18px; margin-top: 32px; }}
  .score-box {{ background: {score_bg}; color: #fff; font-size: 48px; font-weight: bold; text-align: center; padding: 24px; border-radius: 12px; margin: 20px 0; }}
  .score-label {{ font-size: 14px; font-weight: normal; }}
  .not-ready {{ background: #FEE2E2; color: #EF4444; padding: 12px; border-radius: 8px; text-align: center; margin: 12px 0; font-weight: 600; }}
  .category {{ width: 100%; padding: 8px 0; border-bottom: 1px solid #E5E7EB; }}
  .category-name {{ font-weight: 500; }}
  .category-score {{ font-weight: 600; text-align: right; }}
  .categories-table {{ width: 100%; border-collapse: collapse; }}
  .categories-table td {{ padding: 8px 0; border-bottom: 1px solid #E5E7EB; }}
  .categories-table .score-col {{ text-align: right; font-weight: 600; }}
  .issue {{ border-left: 4px solid {issue_color}; padding: 12px 16px; margin: 12px 0; background: #F9FAFB; border-radius: 0 8px 8px 0; }}
  .issue-title {{ font-weight: 600; margin-bottom: 4px; }}
  .severity {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 12px; font-weight: 600; color: #fff; }}
  .severity-critical {{ background: #EF4444; }}
  .severity-serious {{ background: #F97316; }}
  .severity-warning {{ background: #F59E0B; }}
  .severity-advice {{ background: #6B7280; }}
  .quote {{ font-style: italic; color: #6B7280; margin: 8px 0; padding: 8px; background: #F3F4F6; border-radius: 4px; font-family: 'JetBrains Mono', monospace; font-size: 13px; }}
  .recommendation {{ color: #059669; margin-top: 8px; }}
  .standard-ref {{ color: #2563EB; font-size: 13px; }}
  .footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid #E5E7EB; color: #9CA3AF; font-size: 12px; text-align: center; }}
</style>
</head>
<body>
<h1>TZ Analyzer — Отчёт об анализе</h1>
<p><strong>Файл:</strong> {filename}</p>
<p><strong>Режим:</strong> {mode}</p>
<p><strong>Дата:</strong> {date}</p>

<div class="score-box">
  {total_score}<br>
  <span class="score-label">из 100 баллов</span>
</div>

{not_ready_html}

<h2>Оценки по категориям</h2>
<table class="categories-table">
{categories_html}
</table>

<h2>Замечания ({issues_count})</h2>
{issues_html}

<div class="footer">
  Сгенерировано TZ Analyzer MVP v0.1.0
</div>
</body>
</html>
"""

SEVERITY_COLORS = {
    "critical": "#EF4444",
    "serious": "#F97316",
    "warning": "#F59E0B",
    "advice": "#6B7280",
}

SEVERITY_LABELS = {
    "critical": "Критично",
    "serious": "Серьёзно",
    "warning": "Замечание",
    "advice": "Совет",
}

AGENT_LABELS = {
    "structural": "Структура",
    "terminological": "Терминология",
    "logical": "Логика",
    "completeness": "Полнота",
    "scientific": "Научность",
}


@router.get("/api/export/expert-evaluation/xlsx")
async def export_expert_evaluation_xlsx(
    expert_name: str | None = Query(None, description="Имя эксперта — проставляется в колонку 'Эксперт'"),
    expert_comment: str | None = Query(None, description="Комментарий эксперта — проставляется во все строки"),
):
    """Export expert evaluation template pre-filled with AI scores for all completed analyses."""
    async with async_session() as db:
        result = await db.execute(
            select(Analysis)
            .options(selectinload(Analysis.agent_results))
            .where(Analysis.status == "completed")
            .order_by(Analysis.created_at.desc())
        )
        analyses = result.scalars().all()

    # Maps criterion column → (max_points, agent_name_that_sources_score)
    # Agent scores are 0-100; scaled to max_points
    CRITERIA = [
        ("Стратегическая\nрелевантность\n(20%)", 20, "logical"),
        ("Цель и задачи\n(10%)",                  10, "structural"),
        ("Научная\nновизна\n(15%)",                15, "scientific"),
        ("Практическая\nприменимость\n(20%)",      20, "completeness"),
        ("Ожидаемые\nрезультаты\n(15%)",           15, "scientific"),
        ("Соц-экономический\nэффект\n(10%)",       10, "completeness"),
        ("Реализуемость\n(10%)",                   10, "logical"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Экспертная оценка"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    score_fill = PatternFill("solid", fgColor="D6E4F0")
    auto_fill = PatternFill("solid", fgColor="E8F5E9")  # light green = auto-filled

    COLUMNS = (
        [("№", 5), ("Название ТЗ", 30), ("Организация", 20), ("Эксперт", 20)]
        + [(c[0], 14) for c in CRITERIA]
        + [("Итоговый\nбалл", 12), ("Комментарий эксперта", 38)]
    )

    ws.row_dimensions[1].height = 65

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_num, analysis in enumerate(analyses, start=2):
        ws.row_dimensions[row_num].height = 30

        # Build agent score lookup
        agent_scores: dict[str, float] = {}
        for ar in analysis.agent_results:
            if ar.score is not None:
                agent_scores[ar.agent_name] = ar.score

        # Fixed columns: №, Название ТЗ, Организация, Эксперт
        fixed = [row_num - 1, analysis.filename or "Текст", "", expert_name or ""]
        for col_idx, value in enumerate(fixed, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 2 else center_align

        # Criteria score columns (cols 5..5+len(CRITERIA)-1)
        for i, (_, max_pts, agent_name) in enumerate(CRITERIA):
            col_idx = 5 + i
            agent_score = agent_scores.get(agent_name)
            if agent_score is not None:
                # Scale: agent 0-100 → 0-max_pts, rounded to 1 decimal
                value = round(agent_score / 100 * max_pts, 1)
                fill = auto_fill
            else:
                value = None
                fill = score_fill
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align

        # Итоговый балл: SUM of criteria cols
        score_col_start = get_column_letter(5)
        score_col_end = get_column_letter(5 + len(CRITERIA) - 1)
        total_cell = ws.cell(
            row=row_num,
            column=5 + len(CRITERIA),
            value=f"=SUM({score_col_start}{row_num}:{score_col_end}{row_num})",
        )
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        total_cell.alignment = center_align

        # Comment column — pre-fill if expert_comment was provided
        comment_cell = ws.cell(row=row_num, column=5 + len(CRITERIA) + 1, value=expert_comment or "")
        comment_cell.border = thin_border
        comment_cell.alignment = left_align

    # Legend row
    legend_row = len(analyses) + 3
    ws.cell(row=legend_row, column=1, value="🟩 — рассчитано автоматически на основе AI-анализа, 🟦 — заполните вручную")
    ws.cell(row=legend_row, column=1).font = Font(italic=True, size=9, color="555555")

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="expert_evaluation.xlsx"'},
    )


async def _render_report_html(analysis_id: str) -> tuple[str, "Analysis"]:
    """Load analysis and render the report template to HTML. Shared by /pdf and /html."""
    async with async_session() as db:
        result = await db.execute(
            select(Analysis)
            .options(selectinload(Analysis.agent_results), selectinload(Analysis.issues))
            .where(Analysis.id == analysis_id)
        )
        analysis = result.scalar_one_or_none()
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")

        score = analysis.total_score or 0
        score_bg = "#10B981" if score >= 70 else "#F59E0B" if score >= 40 else "#EF4444"

        not_ready_html = ""
        if analysis.not_ready:
            not_ready_html = '<div class="not-ready">⚠ Документ не готов к утверждению</div>'

        # Categories — table rows (xhtml2pdf doesn't support flexbox)
        categories_html = ""
        for ar in analysis.agent_results:
            label = AGENT_LABELS.get(ar.agent_name, ar.agent_name)
            s = ar.score or 0
            color = "#10B981" if s >= 70 else "#F59E0B" if s >= 40 else "#EF4444"
            categories_html += (
                f'<tr><td class="category-name">{label}</td>'
                f'<td class="score-col" style="color:{color}">{s:.0f}/100</td></tr>\n'
            )

        # Issues
        issues_html = ""
        sorted_issues = sorted(analysis.issues, key=lambda i: (
            {"critical": 0, "serious": 1, "warning": 2, "advice": 3}.get(i.severity, 4)
        ))
        for issue in sorted_issues:
            color = SEVERITY_COLORS.get(issue.severity, "#6B7280")
            label = SEVERITY_LABELS.get(issue.severity, issue.severity)
            quote_html = f'<div class="quote">{issue.document_quote}</div>' if issue.document_quote else ""
            ref_html = f'<div class="standard-ref">{issue.standard_reference}</div>' if issue.standard_reference else ""
            issues_html += f"""
<div class="issue" style="border-left-color: {color}">
  <span class="severity severity-{issue.severity}">{label}</span>
  <div class="issue-title">{issue.title}</div>
  <div>{issue.description}</div>
  {quote_html}
  {ref_html}
  <div class="recommendation">💡 {issue.recommendation}</div>
</div>
"""

        html = REPORT_TEMPLATE.format(
            filename=analysis.filename or "Текст",
            mode="Полный" if analysis.mode == "full" else "Быстрый",
            date=analysis.created_at.strftime("%d.%m.%Y %H:%M") if analysis.created_at else "",
            total_score=f"{score:.0f}",
            score_bg=score_bg,
            not_ready_html=not_ready_html,
            categories_html=categories_html,
            issues_count=len(sorted_issues),
            issues_html=issues_html,
            issue_color="#E5E7EB",
        )

        return html, analysis


@router.get("/api/export/{analysis_id}/pdf")
async def export_pdf(analysis_id: str):
    """Render the analysis report as a real PDF file."""
    html, analysis = await _render_report_html(analysis_id)

    if not _PDF_AVAILABLE:
        logger.warning("xhtml2pdf unavailable; returning HTML instead of PDF")
        return HTMLResponse(content=html)

    buf = io.BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html,
        dest=buf,
        encoding="utf-8",
    )
    if pisa_status.err:
        logger.error("PDF generation failed for %s: %s", analysis_id, pisa_status.err)
        raise HTTPException(status_code=500, detail="Не удалось сгенерировать PDF")

    buf.seek(0)
    safe_id = analysis_id[:8]
    filename = f"analysis_{safe_id}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/export/{analysis_id}/html")
async def export_html(analysis_id: str):
    """Return the analysis report as HTML (for in-browser preview / debugging)."""
    html, _ = await _render_report_html(analysis_id)
    return HTMLResponse(content=html)


@router.get("/api/export/{analysis_id}/xlsx")
async def export_xlsx(analysis_id: str):
    async with async_session() as db:
        result = await db.execute(
            select(Analysis)
            .options(selectinload(Analysis.agent_results), selectinload(Analysis.issues))
            .where(Analysis.id == analysis_id)
        )
        analysis = result.scalar_one_or_none()
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")

        data = {
            "filename": analysis.filename,
            "created_at": analysis.created_at,
            "mode": analysis.mode,
            "total_score": analysis.total_score,
            "status": analysis.status,
            "summary": analysis.summary,
            "agent_results": [
                {
                    "agent_name": ar.agent_name,
                    "score": ar.score,
                    "weight": ar.weight,
                    "status": ar.status,
                }
                for ar in analysis.agent_results
            ],
            "issues": [
                {
                    "severity": issue.severity,
                    "title": issue.title,
                    "agent_name": issue.agent_name,
                    "description": issue.description,
                    "recommendation": issue.recommendation,
                    "penalty": issue.penalty,
                }
                for issue in sorted(analysis.issues, key=lambda i: (
                    {"critical": 0, "serious": 1, "warning": 2, "advice": 3}.get(i.severity, 4)
                ))
            ],
        }

        xlsx_bytes = export_single_analysis(data)
        filename = f"analysis_{analysis_id[:8]}.xlsx"

        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@router.get("/api/export/history/xlsx")
async def export_history_xlsx(
    search: str | None = Query(None),
    folder_id: str | None = Query(None),
    min_score: float | None = Query(None, ge=0, le=100),
    max_score: float | None = Query(None, ge=0, le=100),
):
    async with async_session() as db:
        query = select(Analysis)

        if search:
            query = query.where(Analysis.filename.ilike(f"%{search}%"))
        if folder_id:
            query = query.where(Analysis.folder_id == folder_id)
        if min_score is not None:
            query = query.where(Analysis.total_score >= min_score)
        if max_score is not None:
            query = query.where(Analysis.total_score <= max_score)

        query = query.order_by(Analysis.created_at.desc())
        result = await db.execute(query)
        analyses = result.scalars().all()

        items = []
        for a in analyses:
            issue_count_result = await db.execute(
                select(func.count()).select_from(Issue).where(Issue.analysis_id == a.id)
            )
            issues_count = issue_count_result.scalar() or 0

            critical_count_result = await db.execute(
                select(func.count()).select_from(Issue).where(
                    Issue.analysis_id == a.id, Issue.severity == "critical"
                )
            )
            critical_count = critical_count_result.scalar() or 0

            items.append({
                "filename": a.filename,
                "created_at": a.created_at,
                "mode": a.mode,
                "total_score": a.total_score,
                "issues_count": issues_count,
                "critical_count": critical_count,
                "status": a.status,
            })

        xlsx_bytes = export_history(items)

        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="history.xlsx"'},
        )
