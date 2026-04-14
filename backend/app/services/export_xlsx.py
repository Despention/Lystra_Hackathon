import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def _apply_borders(ws, start_row: int, end_row: int, col_count: int):
    """Apply borders to data cells."""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_single_analysis(analysis_data: dict) -> bytes:
    """Create an XLSX workbook for a single analysis with 3 sheets."""
    wb = Workbook()

    # --- Sheet 1: Overview ---
    ws_overview = wb.active
    ws_overview.title = "Обзор"

    overview_headers = ["Параметр", "Значение"]
    ws_overview.append(overview_headers)
    _style_header(ws_overview, 1, len(overview_headers))

    rows = [
        ("Файл", analysis_data.get("filename", "Текст")),
        ("Дата", _format_date(analysis_data.get("created_at"))),
        ("Режим", "Полный" if analysis_data.get("mode") == "full" else "Быстрый"),
        ("Балл", analysis_data.get("total_score", "—")),
        ("Статус", analysis_data.get("status", "—")),
        ("Резюме", analysis_data.get("summary", "—")),
    ]
    for label, value in rows:
        ws_overview.append([label, value])
    _apply_borders(ws_overview, 2, len(rows) + 1, 2)
    _auto_width(ws_overview)

    # --- Sheet 2: Agents ---
    ws_agents = wb.create_sheet("Агенты")

    agent_headers = ["Агент", "Балл", "Вес", "Статус"]
    ws_agents.append(agent_headers)
    _style_header(ws_agents, 1, len(agent_headers))

    agent_results = analysis_data.get("agent_results", [])
    for idx, ar in enumerate(agent_results, start=2):
        ws_agents.append([
            ar.get("agent_name", ""),
            ar.get("score", "—"),
            ar.get("weight", "—"),
            ar.get("status", "—"),
        ])
    _apply_borders(ws_agents, 2, len(agent_results) + 1, len(agent_headers))
    _auto_width(ws_agents)

    # --- Sheet 3: Issues ---
    ws_issues = wb.create_sheet("Замечания")

    issue_headers = ["Серьёзность", "Заголовок", "Агент", "Описание", "Рекомендация", "Штраф"]
    ws_issues.append(issue_headers)
    _style_header(ws_issues, 1, len(issue_headers))

    issues = analysis_data.get("issues", [])
    for idx, issue in enumerate(issues, start=2):
        ws_issues.append([
            issue.get("severity", ""),
            issue.get("title", ""),
            issue.get("agent_name", ""),
            issue.get("description", ""),
            issue.get("recommendation", ""),
            issue.get("penalty", 0),
        ])
    _apply_borders(ws_issues, 2, len(issues) + 1, len(issue_headers))
    _auto_width(ws_issues)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_history(analyses: list[dict]) -> bytes:
    """Create an XLSX workbook for analysis history."""
    wb = Workbook()
    ws = wb.active
    ws.title = "История анализов"

    headers = ["Файл", "Дата", "Режим", "Балл", "Замечаний", "Критических", "Статус"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for idx, a in enumerate(analyses, start=2):
        ws.append([
            a.get("filename", "Текст"),
            _format_date(a.get("created_at")),
            "Полный" if a.get("mode") == "full" else "Быстрый",
            a.get("total_score", "—"),
            a.get("issues_count", 0),
            a.get("critical_count", 0),
            a.get("status", "—"),
        ])
    _apply_borders(ws, 2, len(analyses) + 1, len(headers))
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _format_date(value) -> str:
    """Format a datetime or string to a readable date string."""
    if value is None:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value)
            return dt.strftime("%d.%m.%Y %H:%M")
        except (ValueError, TypeError):
            return value
    return str(value)
